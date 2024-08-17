import tkinter as tk
from tkinter import messagebox, filedialog, ttk, Toplevel, Menu
import os
import sys
import json
from openpyxl import load_workbook
from num2words import num2words
from datetime import datetime
import tempfile
import win32api
import win32print

def resource_path(relative_path):
    """ Get absolute path to resource, works for both dev and PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Config file for persistent storage of company name
config_file = resource_path("config.json")

def load_config():
    if os.path.exists(config_file):
        with open(config_file, 'r') as file:
            return json.load(file)
    else:
        return {"company_name": ""}

def save_config(config):
    with open(config_file, 'w') as file:
        json.dump(config, file)

config = load_config()
master_company_name = config.get("company_name", "")

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Invoice Generator By KMSOFT")
        self.root.geometry("800x600")

        self.font_large = ("Arial", 12)

         # Menu Bar
        menubar = Menu(root)
        root.config(menu=menubar)
        
        # About Menu
        about_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="About", menu=about_menu)
        about_menu.add_command(label="About the Developer", command=self.show_about_info)

        # Company Name
        self.company_label = tk.Label(root, text="Company Name:", font=self.font_large)
        self.company_label.grid(row=0, column=0, padx=10, pady=10)
        self.company_name_var = tk.StringVar(value=master_company_name)
        self.company_entry = tk.Entry(root, textvariable=self.company_name_var, font=self.font_large)
        self.company_entry.grid(row=0, column=1, padx=10, pady=10)

        # Consignee Name
        self.consignee_label = tk.Label(root, text="Consignee Name:", font=self.font_large)
        self.consignee_label.grid(row=1, column=0, padx=10, pady=10)
        self.consignee_name_var = tk.StringVar()
        self.consignee_entry = tk.Entry(root, textvariable=self.consignee_name_var, font=self.font_large)
        self.consignee_entry.grid(row=1, column=1, padx=10, pady=10)

        # Consignee Address
        self.address_label = tk.Label(root, text="Consignee Address:", font=self.font_large)
        self.address_label.grid(row=2, column=0, padx=10, pady=10)
        self.address_var = tk.StringVar()
        self.address_entry = tk.Entry(root, textvariable=self.address_var, font=self.font_large)
        self.address_entry.grid(row=2, column=1, padx=10, pady=10)

        # Invoice Number
        self.invoice_number_label = tk.Label(root, text="Invoice Number:", font=self.font_large)
        self.invoice_number_label.grid(row=3, column=0, padx=10, pady=10)
        self.invoice_number_var = tk.StringVar(value=self.generate_invoice_number())
        self.invoice_number_entry = tk.Entry(root, textvariable=self.invoice_number_var, font=self.font_large)
        self.invoice_number_entry.grid(row=3, column=1, padx=10, pady=10)

        # Invoice Date
        self.invoice_date_label = tk.Label(root, text="Invoice Date:", font=self.font_large)
        self.invoice_date_label.grid(row=4, column=0, padx=10, pady=10)
        self.invoice_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.invoice_date_entry = tk.Entry(root, textvariable=self.invoice_date_var, font=self.font_large)
        self.invoice_date_entry.grid(row=4, column=1, padx=10, pady=10)

        # Item Details: Description, Quantity, Rate
        self.item_frame = tk.Frame(root)
        self.item_frame.grid(row=5, column=0, columnspan=2, pady=10)

        tk.Label(self.item_frame, text="Description", font=self.font_large).grid(row=0, column=0, padx=5)
        tk.Label(self.item_frame, text="Quantity", font=self.font_large).grid(row=0, column=1, padx=5)
        tk.Label(self.item_frame, text="Rate", font=self.font_large).grid(row=0, column=2, padx=5)

        self.description_var = tk.StringVar()
        self.quantity_var = tk.DoubleVar()
        self.rate_var = tk.DoubleVar()

        self.description_entry = tk.Entry(self.item_frame, textvariable=self.description_var, font=self.font_large)
        self.description_entry.grid(row=1, column=0, padx=5)

        self.quantity_entry = tk.Entry(self.item_frame, textvariable=self.quantity_var, font=self.font_large)
        self.quantity_entry.grid(row=1, column=1, padx=5)

        self.rate_entry = tk.Entry(self.item_frame, textvariable=self.rate_var, font=self.font_large)
        self.rate_entry.grid(row=1, column=2, padx=5)

        self.items = []
        self.add_item_button = tk.Button(self.item_frame, text="Add Item", font=self.font_large, command=self.add_item)
        self.add_item_button.grid(row=1, column=3, padx=10)

        # Save Button
        self.save_button = tk.Button(root, text="Save Company Name", font=self.font_large, command=self.save_company_name)
        self.save_button.grid(row=6, column=0, pady=20)

        # Generate Invoice Button
        self.generate_button = tk.Button(root, text="Generate Invoice", font=self.font_large, command=self.generate_invoice)
        self.generate_button.grid(row=6, column=1, pady=20)

        # Preview Invoice Button
        self.preview_button = tk.Button(root, text="Preview Invoice", font=self.font_large, command=self.preview_invoice)
        self.preview_button.grid(row=7, column=0, pady=20)

        # Print Invoice Button
        self.print_button = tk.Button(root, text="Print Invoice", font=self.font_large, command=self.print_invoice)
        self.print_button.grid(row=7, column=1, pady=20)

        # Clear All Button
        self.clear_button = tk.Button(root, text="Clear All", font=self.font_large, command=self.clear_all)
        self.clear_button.grid(row=8, column=0, pady=20)

        # Search Invoice Button
        self.search_button = tk.Button(root, text="Search Invoice", font=self.font_large, command=self.search_invoice)
        self.search_button.grid(row=8, column=1, pady=20)

    def show_about_info(self):
        about_text = (
            "Developer: Krish Maheshwari\n\n"
            "Full Stack Developer working on exciting projects and continuously exploring new technologies.\n\n"
            "Contact Details:\n"
            "Email: krishm.km17@gmail.com\n"
            "GitHub: https://github.com/krish2958\n\n"
            "© 2024 Krish Maheshwari & KMSOFT18\n"
            "Jai Shree Krishna"
        )
        messagebox.showinfo("About the Developer", about_text)

    def generate_invoice_number(self):
        # Auto-generate invoice number logic
        return "INV-" + datetime.now().strftime("%Y%m%d-%H%M%S")

    def save_company_name(self):
        config["company_name"] = self.company_name_var.get()
        save_config(config)
        messagebox.showinfo("Success", "Company name saved successfully!")

    def add_item(self):
        item = {
            "description": self.description_var.get(),
            "quantity": self.quantity_var.get(),
            "rate": self.rate_var.get()
        }
        self.items.append(item)
        self.description_var.set("")
        self.quantity_var.set(0)
        self.rate_var.set(0)
        messagebox.showinfo("Item Added", f"Item '{item['description']}' added successfully!")

    def generate_invoice(self):
        try:
            wb = load_workbook(resource_path('template/invoice_template.xlsx'))
            ws = wb.active

            # Insert Company Name
            ws['F5'] = self.company_name_var.get()

            # Insert Consignee Details
            ws['E11'] = f"{self.consignee_name_var.get()}"
            ws['E12'] = f"{self.address_var.get()}"

            # Insert Invoice Details
            ws['G9'] = f"{self.invoice_number_var.get()}"
            ws['K9'] = f"{self.invoice_date_var.get()}"

            # Insert Items
            total_value = 0
            start_row = 19
            for i, item in enumerate(self.items, start=1):
                ws[f'D{start_row + i - 1}'] = i
                ws[f'E{start_row + i - 1}'] = item['description']
                ws[f'J{start_row + i - 1}'] = 1  # Assuming size as 1
                ws[f'K{start_row + i - 1}'] = item['quantity']
                ws[f'L{start_row + i - 1}'] = item['rate']
                amount = item['quantity'] * item['rate']
                ws[f'M{start_row + i - 1}'] = amount
                total_value += amount

            # Calculate Office Charge and Grand Total
            office_charge = total_value * 0.18
            grand_total = total_value + office_charge

            ws['M34'] = total_value
            ws['M35'] = office_charge
            ws['M36'] = grand_total
            ws['M37'] = grand_total

            # Insert Total in Words
            ws['H38'] = "Rupees " + num2words(grand_total, to='cardinal', lang='en_IN').title() + " Only"

            # Save invoice
            invoice_filename = f"Invoice_{self.consignee_name_var.get()}_{self.invoice_date_var.get()}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=invoice_filename)
            if save_path:
                wb.save(save_path)
                messagebox.showinfo("Success", f"Invoice saved as {save_path}")
            else:
                messagebox.showwarning("Save Cancelled", "Invoice save was cancelled.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def preview_invoice(self):
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Invoice Preview")
        preview_window.geometry("600x400")  # Increase size for better preview

        # Create a table-like display for the preview
        for i, item in enumerate(self.items, start=1):
            tk.Label(preview_window, text=f"{i}", font=self.font_large, borderwidth=1, relief="solid", width=5).grid(row=i, column=0)
            tk.Label(preview_window, text=item['description'], font=self.font_large, borderwidth=1, relief="solid", width=20).grid(row=i, column=1)
            tk.Label(preview_window, text=f"{item['quantity']}", font=self.font_large, borderwidth=1, relief="solid", width=10).grid(row=i, column=2)
            tk.Label(preview_window, text=f"{item['rate']}", font=self.font_large, borderwidth=1, relief="solid", width=10).grid(row=i, column=3)
            tk.Label(preview_window, text=f"{item['quantity'] * item['rate']}", font=self.font_large, borderwidth=1, relief="solid", width=10).grid(row=i, column=4)

        # Display total value, office charge, and grand total
        total_value = sum(item['quantity'] * item['rate'] for item in self.items)
        office_charge = total_value * 0.18
        grand_total = total_value + office_charge

        tk.Label(preview_window, text="Total Value:", font=self.font_large).grid(row=len(self.items) + 1, column=3)
        tk.Label(preview_window, text=f"{total_value}", font=self.font_large).grid(row=len(self.items) + 1, column=4)

        tk.Label(preview_window, text="Office Charge (18%):", font=self.font_large).grid(row=len(self.items) + 2, column=3)
        tk.Label(preview_window, text=f"{office_charge}", font=self.font_large).grid(row=len(self.items) + 2, column=4)

        tk.Label(preview_window, text="Grand Total:", font=self.font_large).grid(row=len(self.items) + 3, column=3)
        tk.Label(preview_window, text=f"{grand_total}", font=self.font_large).grid(row=len(self.items) + 3, column=4)

    def print_invoice(self):
        try:
            wb = load_workbook('invoice_template.xlsx')
            ws = wb.active

            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()

            # Open printer dialog
            printer_name = win32print.GetDefaultPrinter()
            win32api.ShellExecute(0, "print", temp_file.name, f'/d:"{printer_name}"', ".", 0)

            messagebox.showinfo("Success", "Invoice sent to printer.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while printing: {str(e)}")

    def search_invoice(self):
        search_window = tk.Toplevel(self.root)
        search_window.title("Search Invoice")
        search_window.geometry("400x200")  # Increased size for better UX

        tk.Label(search_window, text="Search by Invoice No. or Consignee Name", font=self.font_large).pack(pady=10)

        search_var = tk.StringVar()
        search_entry = tk.Entry(search_window, textvariable=search_var, font=self.font_large, width=30)
        search_entry.pack(pady=10)

        def perform_search():
            search_term = search_var.get().lower()
            for root_dir, dirs,  files in os.walk(os.getcwd()):
                for file in files:
                    if file.endswith(".xlsx"):
                        if search_term in file.lower():
                            found_invoice = os.path.join(root_dir, file)
                            self.show_invoice_found_popup(found_invoice)
                            return
            messagebox.showwarning("No Match", "No matching invoice found.")

        tk.Button(search_window, text="Search", font=self.font_large, command=perform_search).pack(pady=10)
    def show_invoice_found_popup(self, file):
        # Create a new top-level window
        popup = Toplevel(self.root)
        popup.title("Invoice Found")
        popup.geometry("900x150")  # Adjust size as needed

        # Create a label to show the found invoice path
        label = tk.Label(popup, text=f"Invoice found: {file}")
        label.pack(pady=10)

        # Create an "Open Invoice" button
        open_button = tk.Button(popup, text="Open Invoice", command=lambda: self.open_invoice(file), font=('Arial', 12))
        open_button.pack(pady=10)

        # Add a close button to close the popup
        close_button = tk.Button(popup, text="Close", command=popup.destroy, font=('Arial', 12))
        close_button.pack(pady=10)

    def open_invoice(self, file):
        if os.path.exists(file):
            os.startfile(file)
        else:
            messagebox.showerror("Error", f"File not found: {file}")

    def clear_all(self):
        self.consignee_name_var.set("")
        self.address_var.set("")

if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()
